from openpyxl import load_workbook


def ejecutar_accion(path, instruccion):
    wb = load_workbook(path)
    ws = wb.active

    accion = instruccion.get("action")

    # =========================
    # VACIAR UNA SOLA COLUMNA
    # =========================
    if accion == "vaciar_columna":
        columna = (instruccion.get("column") or "").strip().upper()

        if not columna.isalpha():
            raise ValueError("Columna inválida (usa letras: A, B, C... o AA, AB...)")

        for fila in range(2, ws.max_row + 1):  # desde fila 2 para no borrar encabezado
            ws[f"{columna}{fila}"].value = None

    # =========================
    # UNIR NOMBRES
    # =========================
    elif accion == "unir_nombres":
        col_nombre = (instruccion.get("col_nombre") or "").strip().upper()
        col_apellido = (instruccion.get("col_apellido") or "").strip().upper()
        col_destino = (instruccion.get("col_destino") or "").strip().upper()

        if not (col_nombre.isalpha() and col_apellido.isalpha() and col_destino.isalpha()):
            raise ValueError("Columnas inválidas. Ejemplo: unir nombres B C en F")

        for fila in range(2, ws.max_row + 1):
            nombre = ws[f"{col_nombre}{fila}"].value
            apellido = ws[f"{col_apellido}{fila}"].value

            nombre = str(nombre).strip() if nombre is not None else ""
            apellido = str(apellido).strip() if apellido is not None else ""

            ws[f"{col_destino}{fila}"].value = f"{nombre} {apellido}".strip()

        # Si quieres, pone encabezado si estaba vacío
        if ws[f"{col_destino}1"].value in (None, ""):
            ws[f"{col_destino}1"].value = "nombres juntos"

    # =========================
    # PROCESAR BASE (opcional)
    # =========================
    elif accion == "procesar_base":
        pass

    else:
        raise ValueError(f"Acción no soportada: {accion}")

    wb.save(path)
